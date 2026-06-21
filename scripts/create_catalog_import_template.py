"""Create an Excel template for large Directus catalog imports.

Usage:
    python scripts/create_catalog_import_template.py
    python scripts/create_catalog_import_template.py --out catalog_import_template.xlsx
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


HEADERS = [
    "id",
    "source_id",
    "source_system",
    "import_batch",
    "status",
    "stock_status",
    "content_status",
    "sort",
    "category",
    "title",
    "model",
    "specs",
    "storage",
    "color",
    "serial",
    "price",
    "price_text",
    "grade",
    "battery",
    "battery_text",
    "meta_battery",
    "warranty",
    "warranty_text",
    "exit",
    "exit_text",
    "availability",
    "short_description",
    "headline",
    "listing_alt",
    "cta_label",
    "has_detail_page",
    "detail_href",
    "tags",
    "passport",
    "trade",
    "admin_note",
    "image_card",
    "image_card_alt",
    "image_main",
    "image_main_label",
    "image_main_alt",
    "image_main_2",
    "image_main_2_label",
    "image_main_2_alt",
    "image_screen",
    "image_screen_label",
    "image_screen_alt",
    "image_body",
    "image_body_label",
    "image_body_alt",
    "image_defect",
    "image_defect_label",
    "image_defect_alt",
    "image_other",
    "image_other_label",
    "image_other_alt",
]

SAMPLE = {
    "id": "iphone-15-pro-256-blue-001",
    "source_id": "stock-001",
    "source_system": "xlsx",
    "import_batch": "2026-06-stock",
    "status": "draft",
    "stock_status": "available",
    "content_status": "review",
    "sort": 100,
    "category": "iphone",
    "title": "iPhone 15 Pro 256 GB Blue",
    "model": "iPhone 15 Pro",
    "specs": "256 GB",
    "storage": "256 GB",
    "color": "Blue Titanium",
    "serial": "IMEI ...4821",
    "price": 89900,
    "price_text": "89 900 ₽",
    "grade": "A",
    "battery": "92%",
    "battery_text": "92%",
    "meta_battery": "92%",
    "warranty": "90 дней",
    "warranty_text": "Гарантия 90 дней",
    "exit": "до 68 000 ₽",
    "exit_text": "Выход до 68 000 ₽",
    "availability": "В наличии в Москве",
    "short_description": "Чистый корпус, полный ISVOI Passport.",
    "headline": "iPhone 15 Pro с понятной историей",
    "listing_alt": "iPhone 15 Pro Blue Titanium",
    "cta_label": "Смотреть паспорт",
    "has_detail_page": True,
    "detail_href": "/device/iphone-15-pro-256-blue-001",
    "tags": "iphone,club",
    "passport": json.dumps({
        "summaryRows": [
            {"label": "Repair", "value": "not opened", "state": "ok"},
            {"label": "Moisture", "value": "no traces", "state": "ok"},
        ],
        "repair": "not opened",
        "water": "no traces",
        "diagnostics": {
            "status": "passed",
            "checklist": [
                {"text": "Face ID works", "state": "ok"},
                {"text": "Battery checked", "state": "ok"},
            ],
        },
        "condition": {
            "gradeText": "grade A",
            "note": "Minor signs of use.",
            "notes": ["Screen without cracks", "Body without dents"],
            "defectPhotoAlt": "Small cosmetic mark close-up",
        },
        "warranty": {
            "duration": "90 days",
            "covered": "Functional faults under Store terms.",
            "notCovered": "Mechanical damage after purchase.",
        },
        "exitPrice": {
            "headline": "up to 68 000 RUB",
            "buyToday": "68 000 RUB",
            "tradeInEstimate": "up to 68 000 RUB",
            "condition": "after inspection",
            "note": "Final amount depends on condition.",
        },
    }, ensure_ascii=False),
    "trade": json.dumps({
        "options": [
            {"value": 42000, "label": "iPhone 13 Pro - 42 000 RUB"},
            {"value": 26000, "label": "iPhone 12 - 26 000 RUB"},
        ],
    }, ensure_ascii=False),
    "admin_note": "Проверить комплект перед публикацией.",
    "image_card": "iphone-15-pro/card.webp",
    "image_card_alt": "iPhone 15 Pro на белом фоне",
    "image_main": "iphone-15-pro/main.webp",
    "image_main_label": "Главный вид",
    "image_main_alt": "iPhone 15 Pro общий вид",
    "image_main_2": "iphone-15-pro/main-2.webp",
    "image_main_2_label": "Второй ракурс",
    "image_main_2_alt": "iPhone 15 Pro второй ракурс",
    "image_screen": "iphone-15-pro/screen.webp",
    "image_screen_label": "Экран",
    "image_screen_alt": "Экран iPhone 15 Pro",
}

LISTS = {
    "status": ["draft", "published", "archived"],
    "stock_status": ["available", "reserved", "sold", "hidden"],
    "content_status": ["needs_content", "needs_photo", "review", "ready"],
    "category": ["iphone", "ipad", "macbook", "watch", "airpods", "other"],
    "has_detail_page": ["TRUE", "FALSE"],
}

NOTES = [
    ("id", "Public slug and primary key. Required for new rows."),
    ("source_id", "Stable external ID. Lets repeated imports update the same row."),
    ("status", "Keep draft until content/photo QA is complete."),
    ("content_status", "Use review after import, ready before publishing."),
    ("passport", "Optional JSON. Importer also writes it to device_passports."),
    ("trade", "Optional JSON with options[]. Importer also writes it to trade_options."),
    ("image_main_2", "Use numbered columns for multiple images with the same role."),
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Create catalog import XLSX template.")
    parser.add_argument("--out", default="directus/catalog_import_template.xlsx", help="Output .xlsx path.")
    return parser.parse_args()


def autosize(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 36)


def main() -> None:
    args = parse_args()
    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "devices"
    sheet.append(HEADERS)
    sheet.append([SAMPLE.get(header, "") for header in HEADERS])
    sheet.freeze_panes = "A2"

    header_fill = PatternFill(fill_type="solid", fgColor="E8F1FF")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for header, values in LISTS.items():
        column = HEADERS.index(header) + 1
        letter = get_column_letter(column)
        validation = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=True)
        sheet.add_data_validation(validation)
        validation.add(f"{letter}2:{letter}5000")

    autosize(sheet)

    notes = workbook.create_sheet("notes")
    notes.append(["column", "note"])
    for row in NOTES:
        notes.append(row)
    for cell in notes[1]:
        cell.font = Font(bold=True)
    autosize(notes)

    workbook.save(out)
    print(f"Created {out}")


if __name__ == "__main__":
    main()
