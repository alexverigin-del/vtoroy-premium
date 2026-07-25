#!/usr/bin/env python3
"""Create the catalog_v3 XLSX workbook used by Directus operators."""

from pathlib import Path
import argparse

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


SHEETS = {
    "products": [
        "sku", "source_id", "slug", "status", "content_status", "product_type",
        "condition", "sale_mode", "brand_slug", "category_slug", "device_model_slug",
        "title", "model", "color", "price", "stock_quantity", "stock_status",
        "warranty", "warranty_text", "completeness", "short_description", "headline",
        "listing_file_id", "listing_alt", "sort",
    ],
    "images": ["product_sku", "file_id", "status", "role", "label", "alt", "sort"],
    "compatibility": ["accessory_sku", "device_model_slug"],
    "passports": [
        "product_sku", "diagnostics_status", "diagnostics_checklist_json",
        "condition_grade_text", "condition_note", "repair", "water",
        "warranty_duration", "warranty_covered", "warranty_not_covered",
    ],
    "trade_options": ["product_sku", "value", "label", "sort", "is_active"],
}


def build(path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="111827")
    for name, columns in SHEETS.items():
        sheet = workbook.create_sheet(name)
        sheet.append(columns)
        for cell in sheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{sheet.cell(1, len(columns)).coordinate}"
        for index, column in enumerate(columns, 1):
            sheet.column_dimensions[sheet.cell(1, index).column_letter].width = min(
                max(len(column) + 2, 14), 34
            )

    products = workbook["products"]
    validations = {
        "D": '"draft,published,archived"',
        "E": '"needs_content,needs_photo,review,ready"',
        "F": '"device,accessory"',
        "G": '"new,used"',
        "H": '"reservation,inquiry,online"',
        "Q": '"available,reserved,sold,hidden"',
    }
    for column, formula in validations.items():
        validation = DataValidation(type="list", formula1=formula, allow_blank=False)
        products.add_data_validation(validation)
        validation.add(f"{column}2:{column}1000")

    workbook["images"].append(
        ["QA-ACCESSORY-UNI-001", "", "draft", "card", "Главное фото", "Описание товара", 10]
    )
    workbook["compatibility"].append(["QA-ACCESSORY-MODEL-001", "samsung-galaxy-s24"])
    workbook["products"].append(
        [
            "QA-ACCESSORY-UNI-001", "qa-import-example", "qa-import-example", "draft",
            "needs_photo", "accessory", "new", "reservation", "other", "cables", "",
            "Тестовый кабель", "USB-C", "Белый", 1490, 10, "available",
            "12 месяцев", "Гарантия 12 месяцев", "Кабель",
            "Пример строки импорта — не публиковать.", "Универсальный аксессуар", "", "", 900,
        ]
    )
    workbook["products"].append(
        [
            "QA-ACCESSORY-MODEL-001", "qa-import-model-example", "qa-import-model-example",
            "draft", "needs_photo", "accessory", "new", "reservation", "samsung", "cases",
            "", "Тестовый чехол Galaxy S24", "Galaxy S24 Case", "Прозрачный", 1990, 5,
            "available", "6 месяцев", "Гарантия 6 месяцев", "Чехол",
            "Пример модельного аксессуара — не публиковать.", "Точная совместимость",
            "", "", 910,
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("output", nargs="?", default="var/catalog_v3_import_template.xlsx")
    args = parser.parse_args()
    destination = Path(args.output).resolve()
    build(destination)
    print(destination)


if __name__ == "__main__":
    main()
