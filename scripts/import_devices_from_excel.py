"""Import / sync devices into Directus from an Excel workbook.

Expected workbook shape:
  - first row contains column headers;
  - one device per row;
  - `id` is required and is used as the public slug / primary key.

Headers may use either Directus snake_case names (`price_text`) or the current
catalog camelCase names (`priceText`). JSON fields (`tags`, `gallery`,
`passport`, `trade`) may be pasted as JSON strings. `tags` also accepts a
comma-separated list.

Usage:
    python scripts/import_devices_from_excel.py --file stock.xlsx --dry-run
    python scripts/import_devices_from_excel.py --file stock.xlsx

Environment:
    DIRECTUS_URL
    DIRECTUS_TOKEN
"""

from __future__ import annotations

import argparse
from datetime import date, datetime, timezone
import json
import mimetypes
import os
import re
import shutil
import subprocess
import sys
from pathlib import Path
from typing import Any
from urllib.parse import quote
from uuid import uuid4

from openpyxl import load_workbook

try:
    from dotenv import load_dotenv
except ImportError:  # optional convenience when running outside the repo env
    load_dotenv = None

try:
    import requests
except ImportError:  # dry-run/template validation can work without requests installed
    requests = None  # type: ignore[assignment]

REQUESTS_HTTP_ERROR = requests.HTTPError if requests is not None else RuntimeError


COLLECTION = "devices"
IMAGES_COLLECTION = "device_images"
PASSPORTS_COLLECTION = "device_passports"
TRADE_OPTIONS_COLLECTION = "trade_options"
MEDIA_ROLES = ("card", "main", "screen", "body", "defect", "other")
ROLE_DEFAULT_SORT = {
    "card": 0,
    "main": 11,
    "screen": 12,
    "body": 13,
    "defect": 14,
    "other": 15,
}
MEDIA_COLUMN_RE = re.compile(
    r"^image_(?P<role>card|main|screen|body|defect|other)(?:_(?P<ordinal>\d+))?(?:_(?P<meta>label|alt))?$",
)

HEADER_MAP = {
    "priceText": "price_text",
    "batteryText": "battery_text",
    "metaBattery": "meta_battery",
    "warrantyText": "warranty_text",
    "exitText": "exit_text",
    "shortDescription": "short_description",
    "listingImage": "listing_image",
    "listingAlt": "listing_alt",
    "ctaLabel": "cta_label",
    "hasDetailPage": "has_detail_page",
    "detailHref": "detail_href",
    "visualClass": "visual_class",
    "stockStatus": "stock_status",
    "contentStatus": "content_status",
    "sourceSystem": "source_system",
    "sourceId": "source_id",
    "importBatch": "import_batch",
    "importedAt": "imported_at",
    "adminNote": "admin_note",
}

JSON_FIELDS = {"tags", "gallery", "passport", "trade"}
INTEGER_FIELDS = {"price", "sort"}
BOOLEAN_FIELDS = {"has_detail_page"}
STOCK_STATUS_ALIASES = {
    "": "available",
    "in_stock": "available",
    "available": "available",
    "reserved": "reserved",
    "sold": "sold",
    "service": "hidden",
    "hidden": "hidden",
}
DEVICE_FIELDS = {
    "id",
    "status",
    "sort",
    "tags",
    "category",
    "title",
    "model",
    "specs",
    "storage",
    "color",
    "serial",
    "price",
    "price_text",
    "grade",
    "battery",
    "battery_text",
    "meta_battery",
    "warranty",
    "warranty_text",
    "exit",
    "exit_text",
    "availability",
    "short_description",
    "headline",
    "listing_image",
    "listing_alt",
    "cta_label",
    "has_detail_page",
    "detail_href",
    "visual_class",
    "gallery",
    "passport",
    "trade",
    "listing_file",
    "created_at",
    "updated_at",
    "stock_status",
    "content_status",
    "source_system",
    "source_id",
    "import_batch",
    "imported_at",
    "admin_note",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import devices from Excel into Directus.")
    parser.add_argument("--file", required=True, help="Path to the .xlsx workbook.")
    parser.add_argument("--sheet", help="Worksheet name. Defaults to the active sheet.")
    parser.add_argument("--assets-root", default=".", help="Base folder for relative image paths in image_* columns.")
    parser.add_argument("--media-folder", default="ISVOI Device Photos", help="Directus Files folder for uploaded product photos.")
    parser.add_argument("--skip-media", action="store_true", help="Import device rows only; ignore image_* columns.")
    parser.add_argument("--replace-media", action="store_true", help="Overwrite existing device_images/listing_file links.")
    parser.add_argument("--source-system", default="xlsx", help="Default devices.source_system for imported rows.")
    parser.add_argument("--import-batch", help="Default devices.import_batch and device_images.import_batch.")
    parser.add_argument("--default-status", default="draft", choices=("draft", "published", "archived"), help="Default devices.status for new/imported rows.")
    parser.add_argument("--default-stock-status", default="available", choices=("available", "reserved", "sold", "hidden", "in_stock"), help="Default devices.stock_status.")
    parser.add_argument("--allow-extra-fields", action="store_true", help="Allow columns that are not known Directus device fields.")
    parser.add_argument("--dry-run", action="store_true", help="Parse and validate only; do not write.")
    return parser.parse_args()


def load_config(require_token: bool = True) -> dict[str, str]:
    """Read Directus connection settings from the environment."""
    if require_token and requests is None:
        raise SystemExit("The `requests` package is required for real Directus imports. Install scripts/requirements.txt.")
    if load_dotenv is not None:
        load_dotenv()
    url = os.environ.get("DIRECTUS_URL", "").strip()
    token = os.environ.get("DIRECTUS_TOKEN", "").strip()
    if not require_token and not url:
        return {"url": "http://directus.local", "token": ""}
    if not url or (require_token and not token):
        raise SystemExit("DIRECTUS_URL and DIRECTUS_TOKEN must be set (see scripts/.env.example).")
    return {"url": url.rstrip("/"), "token": token}


def headers(cfg: dict[str, str]) -> dict[str, str]:
    return {"Authorization": f"Bearer {cfg['token']}", "Content-Type": "application/json"}


def auth_headers(cfg: dict[str, str]) -> dict[str, str]:
    return {"Authorization": f"Bearer {cfg['token']}"}


def normalize_header(value: Any) -> str:
    raw = str(value or "").strip()
    return HEADER_MAP.get(raw, raw)


def media_column(value: str) -> re.Match[str] | None:
    return MEDIA_COLUMN_RE.match(value)


def is_media_field(value: str) -> bool:
    return media_column(value) is not None


def empty(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = str(value).strip().lower()
    return text in {"1", "true", "yes", "y", "да", "истина"}


def parse_stock_status(value: Any) -> str:
    key = str(value or "").strip().lower()
    if key not in STOCK_STATUS_ALIASES:
        raise ValueError(f"Unknown stock_status `{value}`. Use available, reserved, sold or hidden.")
    return STOCK_STATUS_ALIASES[key]


def parse_json_field(field: str, value: Any) -> Any:
    if empty(value):
        return [] if field in {"tags", "gallery"} else {}
    if not isinstance(value, str):
        return value
    text = value.strip()
    if field == "tags" and not text.startswith("["):
        return [part.strip() for part in text.split(",") if part.strip()]
    return json.loads(text)


def coerce(field: str, value: Any) -> Any:
    if empty(value):
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day, tzinfo=timezone.utc).isoformat()
    if is_media_field(field):
        return str(value).strip()
    if field in JSON_FIELDS:
        return parse_json_field(field, value)
    if field in INTEGER_FIELDS:
        return int(value)
    if field in BOOLEAN_FIELDS:
        return parse_bool(value)
    if field == "stock_status":
        return parse_stock_status(value)
    return value


def read_rows(path: str, sheet_name: str | None = None, *, allow_extra_fields: bool = False) -> list[dict[str, Any]]:
    """Read device rows from the workbook into Directus payloads."""
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [normalize_header(value) for value in rows[0]]
    unknown = sorted({
        header for header in headers
        if header and header not in DEVICE_FIELDS and not is_media_field(header)
    })
    if unknown and not allow_extra_fields:
        raise ValueError(
            "Unknown column(s): "
            + ", ".join(unknown)
            + ". Fix the header or pass --allow-extra-fields intentionally.",
        )

    devices: list[dict[str, Any]] = []
    for row_index, row in enumerate(rows[1:], start=2):
        payload: dict[str, Any] = {}
        for header, value in zip(headers, row):
            if not header or empty(value):
                continue
            payload[header] = coerce(header, value)
        if not payload:
            continue
        if "id" not in payload and "source_id" not in payload:
            raise ValueError(f"Row {row_index}: missing `id`. Existing rows may be updated by `source_id`, but new rows need a public slug in `id`.")
        devices.append(payload)
    return devices


def apply_import_defaults(
    rows: list[dict[str, Any]],
    *,
    source_system: str,
    import_batch: str | None,
    default_status: str,
    default_stock_status: str,
) -> list[dict[str, Any]]:
    imported_at = datetime.now(timezone.utc).isoformat()
    normalized_default_stock_status = parse_stock_status(default_stock_status)
    for row in rows:
        row.setdefault("source_system", source_system)
        row.setdefault("content_status", "review" if any(key.startswith("image_") for key in row) else "needs_photo")
        row.setdefault("stock_status", normalized_default_stock_status)
        row.setdefault("status", default_status)
        if import_batch:
            row.setdefault("import_batch", import_batch)
        row["imported_at"] = imported_at
    return rows


def split_media_fields(row: dict[str, Any]) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    """Separate Directus device fields from image_* import helper columns."""
    device = {key: value for key, value in row.items() if not is_media_field(key)}
    grouped: dict[tuple[str, int], dict[str, Any]] = {}
    for key, value in row.items():
        match = media_column(key)
        if not match:
            continue
        role = match.group("role")
        ordinal = int(match.group("ordinal") or "1")
        meta = match.group("meta") or "path"
        grouped.setdefault((role, ordinal), {"role": role, "ordinal": ordinal})[meta] = value

    media: list[dict[str, Any]] = []
    for (role, ordinal), values in sorted(grouped.items(), key=lambda item: (ROLE_DEFAULT_SORT[item[0][0]], item[0][1])):
        image_path = values.get("path")
        if not image_path:
            continue
        sort = ROLE_DEFAULT_SORT[role] if ordinal == 1 else ROLE_DEFAULT_SORT[role] + ((ordinal - 1) * 100)
        media.append({
            "role": role,
            "ordinal": ordinal,
            "path": image_path,
            "label": values.get("label") or ("Card" if role == "card" else role.title()),
            "alt": values.get("alt") or row.get("listing_alt") or row.get("title") or row.get("id") or row.get("source_id"),
            "sort": sort,
        })
    return device, media


def request_json(cfg: dict[str, str], method: str, endpoint: str, payload: dict[str, Any] | None = None) -> Any:
    res = requests.request(
        method,
        f"{cfg['url']}{endpoint}",
        headers=headers(cfg),
        json=payload,
        timeout=60,
    )
    res.raise_for_status()
    body = res.json() if res.text else {}
    return body.get("data")


def get_all(cfg: dict[str, str], endpoint: str) -> list[dict[str, Any]]:
    data = request_json(cfg, "GET", endpoint)
    return data if isinstance(data, list) else []


def build_indexes(cfg: dict[str, str]) -> dict[str, dict[str, dict[str, Any]]]:
    """Load existing media rows once; faster and safer for large imports."""
    files = get_all(cfg, "/files?fields=id,title,filename_download&limit=-1")
    images = get_all(cfg, "/items/device_images?fields=id,device,role,sort,image&limit=-1")
    return {
        "files": {str(row.get("title")): row for row in files if row.get("title")},
        "images": {
            device_image_key(str(row.get("device")), str(row.get("role")), int(row.get("sort") or 0)): row
            for row in images
            if row.get("device") and row.get("role")
        },
    }


def ensure_folder(cfg: dict[str, str], name: str, dry_run: bool) -> str | None:
    endpoint = f"/folders?filter[name][_eq]={quote(name, safe='')}&fields=id,name&limit=1"
    rows = get_all(cfg, endpoint)
    if rows:
        return str(rows[0]["id"])
    if dry_run:
        print(f"[dry-run] create folder {name}")
        return None
    created = request_json(cfg, "POST", "/folders", {"name": name})
    print(f"[create] folder {name}")
    return str(created["id"])


def find_device_by_source(cfg: dict[str, str], device: dict[str, Any]) -> dict[str, Any] | None:
    source_system = str(device.get("source_system") or "").strip()
    source_id = str(device.get("source_id") or "").strip()
    if not source_system or not source_id:
        return None
    endpoint = (
        f"/items/{COLLECTION}"
        f"?filter[source_system][_eq]={quote(source_system, safe='')}"
        f"&filter[source_id][_eq]={quote(source_id, safe='')}"
        "&fields=id,source_system,source_id&limit=1"
    )
    rows = get_all(cfg, endpoint)
    return rows[0] if rows else None


def upsert_device(cfg: dict[str, str], device: dict[str, Any], dry_run: bool) -> str:
    """Create or update one device in Directus."""
    slug = str(device.get("id") or "").strip()
    source_label = f"{device.get('source_system', '')}/{device.get('source_id', '')}".strip("/")
    if dry_run:
        label = slug or f"source:{source_label}"
        print(f"[dry-run] upsert devices/{label} -> {json.dumps(device, ensure_ascii=False)[:180]}")
        return slug

    base = cfg["url"]
    existing_id = ""
    if slug:
        res = requests.get(f"{base}/items/{COLLECTION}/{slug}", headers=headers(cfg), timeout=30)
        if res.status_code == 200:
            existing_id = slug

    if not existing_id:
        source_match = find_device_by_source(cfg, device)
        if source_match:
            existing_id = str(source_match["id"])
            if slug and slug != existing_id:
                raise ValueError(f"source_id {source_label} already belongs to devices/{existing_id}, not {slug}.")

    if existing_id:
        payload = dict(device)
        payload.pop("id", None)
        patched = requests.patch(f"{base}/items/{COLLECTION}/{existing_id}", headers=headers(cfg), json=payload, timeout=30)
        patched.raise_for_status()
        print(f"[update] devices/{existing_id}")
        return existing_id

    if not slug:
        raise ValueError(f"Cannot create new device from source {source_label}: missing public slug in `id`.")
    created = requests.post(f"{base}/items/{COLLECTION}", headers=headers(cfg), json=device, timeout=30)
    created.raise_for_status()
    print(f"[create] devices/{slug}")
    return slug


def device_image_key(device_id: str, role: str, sort: int) -> str:
    return f"{device_id}:{role}:{sort}"


def directus_file_title(device_id: str, role: str, ordinal: int) -> str:
    if ordinal == 1:
        return f"isvoi:{device_id}:{role}:xlsx"
    return f"isvoi:{device_id}:{role}:{ordinal}:xlsx"


def resolve_image_path(assets_root: Path, value: str) -> Path:
    path = Path(value)
    if path.is_absolute():
        return path
    return assets_root / path


def upload_file(
    cfg: dict[str, str],
    indexes: dict[str, dict[str, dict[str, Any]]],
    *,
    device_id: str,
    role: str,
    ordinal: int,
    image_path: Path,
    folder_id: str | None,
    dry_run: bool,
) -> str | None:
    title = directus_file_title(device_id, role, ordinal)
    existing = indexes["files"].get(title)
    if existing:
        print(f"[skip] file {title} -> {existing['id']}")
        return str(existing["id"])

    if not image_path.exists():
        raise FileNotFoundError(f"Missing image for {device_id}/{role}: {image_path}")
    if dry_run:
        print(f"[dry-run] upload {image_path} as {title}")
        return None

    mime = mimetypes.guess_type(image_path.name)[0] or "application/octet-stream"
    data = {"title": title, "description": f"{device_id} {role}"}
    if folder_id:
        data["folder"] = folder_id
    with image_path.open("rb") as fh:
        res = requests.post(
            f"{cfg['url']}/files",
            headers=auth_headers(cfg),
            data=data,
            files={"file": (image_path.name, fh, mime)},
            timeout=120,
        )
    try:
        res.raise_for_status()
    except REQUESTS_HTTP_ERROR as error:
        fallback_id = local_directus_file_insert(
            image_path,
            title=title,
            description=f"{device_id} {role}",
            folder_id=folder_id,
            mime=mime,
        )
        if fallback_id:
            indexes["files"][title] = {"id": fallback_id, "title": title, "filename_download": image_path.name}
            print(f"[upload:fallback] file {title} -> {fallback_id}")
            return fallback_id
        raise RuntimeError(
            f"Upload {image_path.name} failed: {res.status_code} {res.text[:1000]}"
        ) from error
    data = res.json()["data"]
    indexes["files"][title] = data
    print(f"[upload] file {title} -> {data['id']}")
    return str(data["id"])


def sql_literal(value: object) -> str:
    if value is None:
        return "NULL"
    return "'" + str(value).replace("'", "''") + "'"


def local_directus_file_insert(
    path: Path,
    *,
    title: str,
    description: str,
    folder_id: str | None,
    mime: str,
) -> str:
    """Fallback for Beget-local imports when static-token /files upload is forbidden."""
    compose_file = Path("infra/directus-beget/docker-compose.yml")
    uploads_dir = Path("infra/directus-beget/uploads")
    if not compose_file.exists() or not uploads_dir.exists():
        return ""

    existing_id = local_psql_scalar(
        f"""
SELECT id
FROM directus_files
WHERE title = {sql_literal(title)}
ORDER BY created_on DESC
LIMIT 1;
"""
    )
    if existing_id:
        return existing_id

    file_id = str(uuid4())
    filename_disk = f"{file_id}{path.suffix}"
    target = uploads_dir / filename_disk
    shutil.copyfile(path, target)

    filesize = target.stat().st_size
    sql = f"""
INSERT INTO directus_files (
  id, storage, filename_disk, filename_download, title, description, type, folder, filesize, uploaded_on
) VALUES (
  {sql_literal(file_id)}::uuid,
  'local',
  {sql_literal(filename_disk)},
  {sql_literal(path.name)},
  {sql_literal(title)},
  {sql_literal(description)},
  {sql_literal(mime)},
  {sql_literal(folder_id)}::uuid,
  {filesize},
  now()
)
RETURNING id;
"""
    inserted_id = local_psql_scalar(sql)
    if not inserted_id:
        target.unlink(missing_ok=True)
        raise RuntimeError("Local Directus file insert failed: no id returned")
    return inserted_id


def local_psql_scalar(sql: str) -> str:
    db_user = os.environ.get("DB_USER", "isvoi")
    db_name = os.environ.get("DB_DATABASE", "isvoi")
    result = subprocess.run(
        [
            "docker",
            "compose",
            "-f",
            "infra/directus-beget/docker-compose.yml",
            "exec",
            "-T",
            "database",
            "psql",
            "-U",
            db_user,
            "-d",
            db_name,
            "-At",
            "-v",
            "ON_ERROR_STOP=1",
        ],
        input=sql,
        text=True,
        capture_output=True,
        check=False,
    )
    if result.returncode != 0:
        raise RuntimeError(f"Local Directus file insert failed: {result.stderr or result.stdout}")
    for line in result.stdout.strip().splitlines():
        value = line.strip()
        if len(value) == 36 and value.count("-") == 4:
            return value
    return ""


def sync_device_image(
    cfg: dict[str, str],
    indexes: dict[str, dict[str, dict[str, Any]]],
    *,
    device_id: str,
    media: dict[str, Any],
    file_id: str | None,
    dry_run: bool,
    replace: bool,
) -> None:
    role = str(media["role"])
    key = device_image_key(device_id, role, int(media["sort"]))
    existing = indexes["images"].get(key)
    payload = {
        "status": "published",
        "sort": media["sort"],
        "device": device_id,
        "role": role,
        "image": file_id,
        "label": media["label"],
        "alt": media["alt"],
        "shot_status": "approved",
        "source_path": str(media.get("path") or ""),
    }
    if media.get("import_batch"):
        payload["import_batch"] = media["import_batch"]

    if existing and not replace:
        print(f"[skip] device_images {device_id}/{role} sort={media['sort']}")
        return
    if dry_run:
        print(f"[dry-run] {'patch' if existing else 'create'} device_images {device_id}/{role} sort={media['sort']}")
        return
    if existing:
        data = request_json(cfg, "PATCH", f"/items/{IMAGES_COLLECTION}/{existing['id']}", payload)
        indexes["images"][key] = data
        print(f"[patch] device_images {device_id}/{role} sort={media['sort']}")
        return
    data = request_json(cfg, "POST", f"/items/{IMAGES_COLLECTION}", payload)
    indexes["images"][key] = data
    print(f"[create] device_images {device_id}/{role} sort={media['sort']}")


def sync_listing_file(cfg: dict[str, str], device_id: str, file_id: str | None, dry_run: bool, replace: bool) -> None:
    if not file_id:
        if dry_run:
            print(f"[dry-run] patch devices/{device_id}.listing_file after upload")
        return
    if dry_run:
        print(f"[dry-run] patch devices/{device_id}.listing_file = {file_id}")
        return
    existing = request_json(cfg, "GET", f"/items/{COLLECTION}/{device_id}?fields=id,listing_file")
    if existing.get("listing_file") and not replace:
        print(f"[skip] devices/{device_id}.listing_file already set")
        return
    request_json(cfg, "PATCH", f"/items/{COLLECTION}/{device_id}", {"listing_file": file_id})
    print(f"[patch] devices/{device_id}.listing_file")


def passport_payload(passport: Any) -> dict[str, Any]:
    if not isinstance(passport, dict) or not passport:
        return {}
    diagnostics = passport.get("diagnostics") if isinstance(passport.get("diagnostics"), dict) else {}
    condition = passport.get("condition") if isinstance(passport.get("condition"), dict) else {}
    warranty = passport.get("warranty") if isinstance(passport.get("warranty"), dict) else {}
    exit_price = passport.get("exitPrice") if isinstance(passport.get("exitPrice"), dict) else {}
    payload = {
        "repair": passport.get("repair"),
        "water": passport.get("water"),
        "summary_rows": passport.get("summaryRows"),
        "diagnostics_status": diagnostics.get("status"),
        "diagnostics_checklist": diagnostics.get("checklist"),
        "condition_grade_text": condition.get("gradeText"),
        "condition_note": condition.get("note"),
        "condition_notes": condition.get("notes"),
        "defect_photo_alt": condition.get("defectPhotoAlt"),
        "warranty_duration": warranty.get("duration"),
        "warranty_covered": warranty.get("covered"),
        "warranty_not_covered": warranty.get("notCovered"),
        "exit_headline": exit_price.get("headline"),
        "exit_buy_today": exit_price.get("buyToday"),
        "exit_trade_in_estimate": exit_price.get("tradeInEstimate"),
        "exit_condition": exit_price.get("condition"),
        "exit_note": exit_price.get("note"),
    }
    return {key: value for key, value in payload.items() if value is not None}


def sync_passport(cfg: dict[str, str], device_id: str, passport: Any, dry_run: bool) -> None:
    payload = passport_payload(passport)
    if not payload:
        return
    payload["device"] = device_id
    if dry_run:
        print(f"[dry-run] upsert device_passports/{device_id} -> {json.dumps(payload, ensure_ascii=False)[:180]}")
        return
    rows = get_all(
        cfg,
        f"/items/{PASSPORTS_COLLECTION}?filter[device][_eq]={quote(device_id, safe='')}&fields=id&limit=1",
    )
    if rows:
        request_json(cfg, "PATCH", f"/items/{PASSPORTS_COLLECTION}/{rows[0]['id']}", payload)
        print(f"[patch] device_passports/{device_id}")
        return
    request_json(cfg, "POST", f"/items/{PASSPORTS_COLLECTION}", payload)
    print(f"[create] device_passports/{device_id}")


def sync_passport_defect_photo(cfg: dict[str, str], device_id: str, file_id: str | None, dry_run: bool) -> None:
    if not file_id:
        return
    if dry_run:
        print(f"[dry-run] patch device_passports/{device_id}.defect_photo = {file_id}")
        return
    rows = get_all(
        cfg,
        f"/items/{PASSPORTS_COLLECTION}?filter[device][_eq]={quote(device_id, safe='')}&fields=id,defect_photo&limit=1",
    )
    if not rows:
        request_json(cfg, "POST", f"/items/{PASSPORTS_COLLECTION}", {"device": device_id, "defect_photo": file_id})
        print(f"[create] device_passports/{device_id}.defect_photo")
        return
    if rows[0].get("defect_photo"):
        print(f"[skip] device_passports/{device_id}.defect_photo already set")
        return
    request_json(cfg, "PATCH", f"/items/{PASSPORTS_COLLECTION}/{rows[0]['id']}", {"defect_photo": file_id})
    print(f"[patch] device_passports/{device_id}.defect_photo")


def trade_option_payloads(trade: Any) -> list[dict[str, Any]]:
    if not isinstance(trade, dict):
        return []
    options = trade.get("options")
    if not isinstance(options, list):
        return []
    payloads: list[dict[str, Any]] = []
    for index, option in enumerate(options, start=1):
        if not isinstance(option, dict):
            continue
        label = str(option.get("label") or "").strip()
        value = option.get("value")
        if not label and empty(value):
            continue
        payloads.append({
            "sort": index * 10,
            "label": label,
            "value": int(value or 0),
            "is_active": True,
        })
    return payloads


def sync_trade_options(cfg: dict[str, str], device_id: str, trade: Any, dry_run: bool) -> None:
    payloads = trade_option_payloads(trade)
    if not payloads:
        return
    if dry_run:
        print(f"[dry-run] sync {len(payloads)} trade_options for {device_id}")
        return
    existing = get_all(
        cfg,
        f"/items/{TRADE_OPTIONS_COLLECTION}?filter[device][_eq]={quote(device_id, safe='')}&fields=id,sort&limit=-1",
    )
    by_sort = {int(row.get("sort") or 0): row for row in existing}
    active_sorts: set[int] = set()
    for payload in payloads:
        payload["device"] = device_id
        sort = int(payload["sort"])
        active_sorts.add(sort)
        row = by_sort.get(sort)
        if row:
            request_json(cfg, "PATCH", f"/items/{TRADE_OPTIONS_COLLECTION}/{row['id']}", payload)
            print(f"[patch] trade_options/{device_id} sort={sort}")
        else:
            request_json(cfg, "POST", f"/items/{TRADE_OPTIONS_COLLECTION}", payload)
            print(f"[create] trade_options/{device_id} sort={sort}")
    for row in existing:
        sort = int(row.get("sort") or 0)
        if sort not in active_sorts:
            request_json(cfg, "PATCH", f"/items/{TRADE_OPTIONS_COLLECTION}/{row['id']}", {"is_active": False})
            print(f"[archive] trade_options/{device_id} sort={sort}")


def sync_structured_catalog_data(cfg: dict[str, str], device_id: str, device: dict[str, Any], dry_run: bool) -> None:
    if not device_id:
        return
    sync_passport(cfg, device_id, device.get("passport"), dry_run)
    sync_trade_options(cfg, device_id, device.get("trade"), dry_run)


def sync_media(
    cfg: dict[str, str],
    indexes: dict[str, dict[str, dict[str, Any]]],
    *,
    device: dict[str, Any],
    media_rows: list[dict[str, Any]],
    assets_root: Path,
    folder_id: str | None,
    dry_run: bool,
    replace: bool,
) -> None:
    device_id = str(device["id"])
    for media in media_rows:
        role = str(media["role"])
        image_path = resolve_image_path(assets_root, str(media["path"]))
        file_id = upload_file(
            cfg,
            indexes,
            device_id=device_id,
            role=role,
            ordinal=int(media.get("ordinal") or 1),
            image_path=image_path,
            folder_id=folder_id,
            dry_run=dry_run,
        )
        sync_device_image(cfg, indexes, device_id=device_id, media=media, file_id=file_id, dry_run=dry_run, replace=replace)
        if role == "card":
            sync_listing_file(cfg, device_id, file_id, dry_run, replace)
        if role == "defect":
            sync_passport_defect_photo(cfg, device_id, file_id, dry_run)


def main() -> None:
    args = parse_args()
    cfg = load_config(require_token=not args.dry_run)
    rows = apply_import_defaults(
        read_rows(args.file, args.sheet, allow_extra_fields=args.allow_extra_fields),
        source_system=args.source_system,
        import_batch=args.import_batch,
        default_status=args.default_status,
        default_stock_status=args.default_stock_status,
    )
    if not rows:
        raise SystemExit(f"No device rows found in {args.file}")
    print(f"{'[dry-run] ' if args.dry_run else ''}Importing {len(rows)} device(s)")
    assets_root = Path(args.assets_root).resolve()
    parsed = [split_media_fields(row) for row in rows]
    if args.import_batch:
        parsed = [
            (device, [{**media, "import_batch": args.import_batch} for media in media_rows])
            for device, media_rows in parsed
        ]
    indexes = None
    folder_id = None
    if args.dry_run:
        indexes = {"files": {}, "images": {}}
    elif not args.skip_media and any(media for _, media in parsed):
        indexes = build_indexes(cfg)
        folder_id = ensure_folder(cfg, args.media_folder, dry_run=False)
    for device, media_rows in parsed:
        device_id = upsert_device(cfg, device, dry_run=args.dry_run)
        sync_structured_catalog_data(cfg, device_id, device, args.dry_run)
        if not args.skip_media and media_rows:
            if indexes is None:
                indexes = build_indexes(cfg)
            sync_media(
                cfg,
                indexes,
                device={**device, "id": device_id or device.get("id")},
                media_rows=media_rows,
                assets_root=assets_root,
                folder_id=folder_id,
                dry_run=args.dry_run,
                replace=args.replace_media,
            )


if __name__ == "__main__":
    try:
        main()
    except (REQUESTS_HTTP_ERROR, ValueError, FileNotFoundError, json.JSONDecodeError) as exc:
        print(f"Import error: {exc}", file=sys.stderr)
        raise SystemExit(1) from exc
