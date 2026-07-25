#!/usr/bin/env python3
"""Validate or upsert a catalog_v3 workbook into Directus."""

from __future__ import annotations

import argparse
import json
import os
from pathlib import Path
from typing import Any
from urllib.parse import quote, urlencode
from urllib.request import Request, urlopen

from openpyxl import load_workbook

from create_product_catalog_import_template import SHEETS


def rows(sheet: Any) -> list[dict[str, Any]]:
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        return []
    header = [str(value or "").strip() for value in values[0]]
    return [
        {header[index]: value for index, value in enumerate(row) if index < len(header)}
        for row in values[1:]
        if any(value not in (None, "") for value in row)
    ]


def validate(workbook: Any) -> dict[str, list[dict[str, Any]]]:
    errors: list[str] = []
    result: dict[str, list[dict[str, Any]]] = {}
    for sheet_name, expected in SHEETS.items():
        if sheet_name not in workbook.sheetnames:
            errors.append(f"missing sheet: {sheet_name}")
            continue
        actual = [str(cell.value or "").strip() for cell in workbook[sheet_name][1]]
        missing = [column for column in expected if column not in actual]
        if missing:
            errors.append(f"{sheet_name}: missing columns {', '.join(missing)}")
        result[sheet_name] = rows(workbook[sheet_name])

    seen: set[str] = set()
    for index, product in enumerate(result.get("products", []), 2):
        sku = str(product.get("sku") or "").strip()
        if not sku:
            errors.append(f"products:{index}: sku is required")
        if sku in seen:
            errors.append(f"products:{index}: duplicate sku {sku}")
        seen.add(sku)
        product_type = product.get("product_type")
        condition = product.get("condition")
        if product_type not in {"device", "accessory"}:
            errors.append(f"products:{index}: invalid product_type")
        if condition not in {"new", "used"}:
            errors.append(f"products:{index}: invalid condition")
        if product_type == "accessory" and condition != "new":
            errors.append(f"products:{index}: accessories must be new")
        if product.get("status") == "published" and product.get("content_status") != "ready":
            errors.append(f"products:{index}: published product must be ready")

    product_skus = seen
    for sheet_name, key in (
        ("images", "product_sku"),
        ("passports", "product_sku"),
        ("trade_options", "product_sku"),
        ("compatibility", "accessory_sku"),
    ):
        for index, row in enumerate(result.get(sheet_name, []), 2):
            if str(row.get(key) or "").strip() not in product_skus:
                errors.append(f"{sheet_name}:{index}: unknown SKU in {key}")

    if errors:
        raise ValueError("\n".join(errors))
    return result


class Directus:
    def __init__(self, base_url: str, token: str) -> None:
        self.base_url = base_url.rstrip("/")
        self.token = token

    def request(self, method: str, path: str, payload: Any = None) -> Any:
        body = json.dumps(payload, ensure_ascii=False).encode() if payload is not None else None
        request = Request(
            f"{self.base_url}{path}",
            data=body,
            method=method,
            headers={"Authorization": f"Bearer {self.token}", "Content-Type": "application/json"},
        )
        with urlopen(request, timeout=45) as response:
            return json.loads(response.read().decode())

    def lookup(self, collection: str, field: str, value: str) -> str | None:
        query = urlencode({f"filter[{field}][_eq]": value, "fields": "id", "limit": 1})
        data = self.request("GET", f"/items/{collection}?{query}").get("data", [])
        return str(data[0]["id"]) if data else None

    def upsert(self, collection: str, lookup_field: str, lookup_value: str, payload: dict[str, Any]) -> str:
        existing = self.lookup(collection, lookup_field, lookup_value)
        if existing:
            self.request("PATCH", f"/items/{collection}/{quote(existing, safe='')}", payload)
            return existing
        created = self.request("POST", f"/items/{collection}", payload)["data"]
        return str(created["id"])


def clean(row: dict[str, Any]) -> dict[str, Any]:
    return {
        key: value
        for key, value in row.items()
        if value not in (None, "") and not key.endswith("_slug") and key not in {"slug", "product_sku", "accessory_sku", "file_id"}
    }


def run_import(data: dict[str, list[dict[str, Any]]], client: Directus) -> None:
    products_by_sku: dict[str, str] = {}
    for row in data["products"]:
        payload = clean(row)
        payload["id"] = str(row.get("slug") or row["sku"]).strip().lower()
        payload["brand"] = client.lookup("product_brands", "slug", str(row["brand_slug"]))
        payload["category"] = client.lookup("product_categories", "slug", str(row["category_slug"]))
        if row.get("device_model_slug"):
            payload["device_model"] = client.lookup("device_models", "slug", str(row["device_model_slug"]))
        if row.get("listing_file_id"):
            payload["listing_file"] = str(row["listing_file_id"])
        product_id = client.upsert("products", "sku", str(row["sku"]), payload)
        products_by_sku[str(row["sku"])] = product_id

    for row in data["images"]:
        payload = clean(row)
        payload["product"] = products_by_sku[str(row["product_sku"])]
        payload["image"] = str(row["file_id"])
        key = f"{payload['product']}:{int(payload.get('sort', 100))}"
        existing_query = urlencode({
            "filter[product][_eq]": payload["product"],
            "filter[sort][_eq]": int(payload.get("sort", 100)),
            "fields": "id",
            "limit": 1,
        })
        existing = client.request("GET", f"/items/product_images?{existing_query}").get("data", [])
        if existing:
            client.request("PATCH", f"/items/product_images/{existing[0]['id']}", payload)
        else:
            client.request("POST", "/items/product_images", payload)

    for row in data["compatibility"]:
        product_id = products_by_sku[str(row["accessory_sku"])]
        model_id = client.lookup("device_models", "slug", str(row["device_model_slug"]))
        query = urlencode({
            "filter[product][_eq]": product_id,
            "filter[device_models_id][_eq]": model_id or "",
            "fields": "id",
            "limit": 1,
        })
        if not client.request("GET", f"/items/product_compatible_models?{query}").get("data"):
            client.request("POST", "/items/product_compatible_models", {
                "product": product_id, "device_models_id": model_id
            })

    for row in data["passports"]:
        payload = clean(row)
        payload["product"] = products_by_sku[str(row["product_sku"])]
        if row.get("diagnostics_checklist_json"):
            payload["diagnostics_checklist"] = json.loads(str(row["diagnostics_checklist_json"]))
        payload.pop("diagnostics_checklist_json", None)
        client.upsert("device_passports", "product", payload["product"], payload)

    for row in data["trade_options"]:
        payload = clean(row)
        payload["product"] = products_by_sku[str(row["product_sku"])]
        query = urlencode({
            "filter[product][_eq]": payload["product"],
            "filter[sort][_eq]": int(payload.get("sort", 100)),
            "fields": "id",
            "limit": 1,
        })
        existing = client.request("GET", f"/items/trade_options?{query}").get("data", [])
        if existing:
            client.request("PATCH", f"/items/trade_options/{existing[0]['id']}", payload)
        else:
            client.request("POST", "/items/trade_options", payload)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    data = validate(load_workbook(Path(args.workbook), data_only=True))
    summary = {name: len(items) for name, items in data.items()}
    print(json.dumps({"valid": True, "rows": summary}, ensure_ascii=False))
    if args.dry_run:
        return
    base_url = os.getenv("DIRECTUS_URL", "")
    token = os.getenv("DIRECTUS_TOKEN", "")
    if not base_url or not token:
        raise RuntimeError("DIRECTUS_URL and DIRECTUS_TOKEN are required")
    run_import(data, Directus(base_url, token))
    print("catalog_v3 import complete")


if __name__ == "__main__":
    main()
