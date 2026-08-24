#!/usr/bin/env python3
"""Validate or apply a private inventory snapshot and receipt workbook."""

from __future__ import annotations

import argparse
import json
import os
import re
import unicodedata
from dataclasses import asdict, dataclass
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from urllib.error import HTTPError
from urllib.parse import quote, urlencode
from urllib.request import Request, urlopen

from openpyxl import load_workbook


INVENTORY_REQUIRED = {
    "Uuid",
    "Наименование",
    "Код",
    "Артикул",
    "В продаже",
    "Цена закупки",
    "Цена",
    "Остаток",
    "Группа",
    "Описание",
    "Штрих-код",
    "Структура групп",
}
RECEIPT_REQUIRED = {
    "№",
    "Наименование",
    "Категория",
    "Серийный номер",
    "Подкатегория",
    "Количество",
    "Закупка",
    "СуммаЗ",
    "Наценка",
    "Маржинальность",
    "Маржа",
    "Продажа",
    "СуммаП",
}


@dataclass
class Issue:
    severity: str
    code: str
    source_kind: str
    row_number: int | None
    source_id: str | None
    message: str


def text(value: Any) -> str:
    return str(value or "").strip()


def number(value: Any, field: str, row_number: int) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError) as error:
        raise ValueError(f"row {row_number}: {field} must be numeric") from error


def integer(value: Any, field: str, row_number: int) -> int:
    parsed = number(value, field, row_number)
    if parsed != parsed.to_integral_value():
        raise ValueError(f"row {row_number}: {field} must be an integer")
    return int(parsed)


def normalize(value: Any) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[^a-zа-я0-9]+", " ", text(value).lower().replace("ё", "е"))).strip()


def relaxed_name(value: Any) -> str:
    cleaned = re.sub(r"\bby apple\b|\bmodel\b|\breplica\b", " ", text(value), flags=re.I)
    return normalize(cleaned)


def serialized_identity(value: Any) -> str:
    generic_tokens = {"мобильный", "телефон", "смартфон", "5g"}
    tokens = [token for token in normalize(value).split() if token not in generic_tokens]
    if "silver" in tokens or "серебристый" in tokens:
        tokens = [token for token in tokens if token not in {"white", "белый"}]
    return " ".join(tokens)


def extract_serial(value: Any) -> str:
    match = re.search(r"Серийный номер:\s*([^\s,;]+)", text(value), flags=re.I)
    return match.group(1).upper() if match else ""


def masked(value: str) -> str:
    value = text(value)
    return f"***{value[-4:]}" if value else ""


def parse_source_datetime(value: Any) -> str | None:
    raw = text(value)
    if not raw:
        return None
    for fmt in ("%d-%m-%Y, %H:%M:%S", "%d.%m.%Y %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(raw, fmt).replace(tzinfo=timezone.utc).isoformat()
        except ValueError:
            continue
    return None


def parse_source_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raw = text(value)
    if not raw:
        return None
    for fmt in ("%d-%m-%Y", "%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def rows_from_sheet(sheet: Any, header_row: int) -> tuple[list[str], list[tuple[int, dict[str, Any]]]]:
    values = list(sheet.iter_rows(values_only=True))
    if len(values) < header_row:
        return [], []
    header = [text(value) for value in values[header_row - 1]]
    rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, values_row in enumerate(values[header_row:], header_row + 1):
        if not any(value not in (None, "") for value in values_row):
            continue
        rows.append((row_number, {header[index]: value for index, value in enumerate(values_row) if index < len(header)}))
    return header, rows


def parse_inventory(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    header, source_rows = rows_from_sheet(workbook[workbook.sheetnames[0]], 1)
    workbook.close()
    missing = sorted(INVENTORY_REQUIRED - set(header))
    if missing:
        raise ValueError(f"inventory workbook missing columns: {', '.join(missing)}")

    rows: list[dict[str, Any]] = []
    seen: dict[str, set[str]] = {"source_id": set(), "source_sku": set(), "barcode": set()}
    for row_number, row in source_rows:
        source_id = text(row.get("Uuid"))
        source_sku = text(row.get("Код"))
        barcode = text(row.get("Штрих-код"))
        title = text(row.get("Наименование"))
        if not source_id or not source_sku or not title:
            raise ValueError(f"inventory row {row_number}: Uuid, Код and Наименование are required")
        for key, value in (("source_id", source_id), ("source_sku", source_sku), ("barcode", barcode)):
            if value and value in seen[key]:
                raise ValueError(f"inventory row {row_number}: duplicate {key} {value}")
            if value:
                seen[key].add(value)
        quantity = integer(row.get("Остаток"), "Остаток", row_number)
        purchase = number(row.get("Цена закупки"), "Цена закупки", row_number)
        retail = number(row.get("Цена"), "Цена", row_number)
        if quantity < 0 or purchase < 0 or retail < 0:
            raise ValueError(f"inventory row {row_number}: stock and prices must be non-negative")
        serial = extract_serial(row.get("Описание"))
        rows.append(
            {
                "row_number": row_number,
                "source_id": source_id,
                "source_sku": source_sku,
                "source_article": text(row.get("Артикул")),
                "barcode": barcode,
                "source_title": title,
                "source_description": text(row.get("Описание")),
                "source_group": text(row.get("Группа")),
                "source_group_path": text(row.get("Структура групп")),
                "serial_full": serial,
                "quantity": quantity,
                "purchase_price": purchase,
                "retail_price": retail,
                "for_sale": bool(row.get("В продаже")),
                "ownership": text(row.get("Тип собственности")),
                "source_created_at": parse_source_datetime(row.get("Создан")),
                "source_updated_at": parse_source_datetime(row.get("Обновлен")),
                "item_kind": "serialized" if serial else "pooled",
                "condition": "used" if serial else "new",
            }
        )
    return rows


def parse_receipts(path: Path | None) -> list[dict[str, Any]]:
    if path is None:
        return []
    workbook = load_workbook(path, read_only=True, data_only=True)
    header, source_rows = rows_from_sheet(workbook[workbook.sheetnames[0]], 2)
    workbook.close()
    missing = sorted(RECEIPT_REQUIRED - set(header))
    if missing:
        raise ValueError(f"receipts workbook missing columns: {', '.join(missing)}")

    rows: list[dict[str, Any]] = []
    for row_number, row in source_rows:
        if not isinstance(row.get("№"), (int, float)):
            continue
        quantity = integer(row.get("Количество"), "Количество", row_number)
        unit_cost = number(row.get("Закупка"), "Закупка", row_number)
        target_price = number(row.get("Продажа"), "Продажа", row_number)
        total_cost = number(row.get("СуммаЗ"), "СуммаЗ", row_number)
        total_price = number(row.get("СуммаП"), "СуммаП", row_number)
        if quantity <= 0 or unit_cost < 0 or target_price < 0:
            raise ValueError(f"receipts row {row_number}: quantity must be positive and prices non-negative")
        expected_cost = unit_cost * quantity
        expected_total = target_price * quantity
        if abs(expected_cost - total_cost) > Decimal("0.02"):
            raise ValueError(f"receipts row {row_number}: СуммаЗ does not match quantity * cost")
        if abs(expected_total - total_price) > Decimal("0.02"):
            raise ValueError(f"receipts row {row_number}: СуммаП does not match quantity * sale")
        rows.append(
            {
                "row_number": row_number,
                "source_title": text(row.get("Наименование")),
                "source_category": text(row.get("Категория")),
                "source_subcategory": text(row.get("Подкатегория")),
                "received_on": parse_source_date(row.get("Дата поступления")),
                "imei_full": text(row.get("IMEI")),
                "serial_full": text(row.get("Серийный номер")).upper(),
                "quantity": quantity,
                "unit_cost": unit_cost,
                "target_markup": number(row.get("Наценка"), "Наценка", row_number),
                "target_margin": number(row.get("Маржинальность"), "Маржинальность", row_number),
                "target_price": target_price,
                "total_cost": total_cost,
                "total_price": total_price,
                "source_note": text(row.get("Примечание") or row.get("Комментарий") or row.get("Ответственный")),
            }
        )
    return rows


def risk_codes(title: str, purchase_price: Decimal) -> list[str]:
    value = normalize(title)
    risks: list[str] = []
    if "replica" in value or "mini ip17" in value or "pods max" in value:
        risks.append("replica_or_mimic")
    if "airpods" in value and purchase_price < Decimal("10000"):
        risks.append("authenticity_review")
    if "jbl" in value and "flip" in value and purchase_price < Decimal("5000"):
        risks.append("authenticity_review")
    if "ассорт" in value:
        risks.append("variant_ambiguous")
    return risks


def classify_receipt_movement(receipt: dict[str, Any], match: dict[str, Any] | None) -> tuple[str, int, str]:
    note = normalize(receipt.get("source_note"))
    is_central_office = "цо" in note.split() or "центральн" in note
    quantity = int(receipt["quantity"])
    stock = int(match["quantity"]) if match else 0

    if is_central_office and match is None:
        return (
            "central_office",
            quantity,
            "Позиция отсутствует в snapshot магазина и учтена в ЦО по комментарию поступления.",
        )
    if is_central_office and quantity > stock:
        central_office_quantity = quantity - stock
        return (
            "partial_central_office",
            central_office_quantity,
            f"В текущем остатке {stock}; разница {central_office_quantity} ед. учтена в ЦО.",
        )
    if is_central_office:
        return (
            "central_office_inventory_conflict",
            0,
            "Комментарий указывает ЦО, но позиция одновременно присутствует в текущем остатке магазина.",
        )
    if match and stock > 0:
        return "in_store", 0, "Позиция сопоставлена с актуальным остатком магазина."
    if receipt.get("source_note"):
        return (
            "exited_preload",
            0,
            f"Нет в текущем остатке; исходный комментарий: {receipt['source_note']}.",
        )
    return (
        "exited_preload",
        0,
        "Нет в текущем остатке; учтено как продажа или иное выбытие до загрузки в магазин.",
    )


def reconcile(inventory: list[dict[str, Any]], receipts: list[dict[str, Any]]) -> tuple[list[Issue], dict[int, dict[str, Any]]]:
    issues: list[Issue] = []
    by_serial = {row["serial_full"]: row for row in inventory if row["serial_full"]}
    by_name = {relaxed_name(row["source_title"]): row for row in inventory}
    receipt_matches: dict[int, dict[str, Any]] = {}
    identity_by_source: dict[str, str] = {}
    receipt_title_by_source: dict[str, str] = {}

    for receipt in receipts:
        serial = receipt["serial_full"]
        match = by_serial.get(serial) if serial else by_name.get(relaxed_name(receipt["source_title"]))
        movement_status, central_office_quantity, match_note = classify_receipt_movement(receipt, match)
        receipt["movement_status"] = movement_status
        receipt["central_office_quantity"] = central_office_quantity
        receipt["match_note"] = match_note

        if match is None:
            receipt["match_status"] = "not_in_snapshot"
            note = normalize(receipt.get("source_note"))
            has_known_exit_note = any(marker in note for marker in ("обмен", "продан", "продано", "продаж", "выбыл"))
            if movement_status == "exited_preload" and not has_known_exit_note:
                issues.append(
                    Issue(
                        "warning",
                        "receipt_exit_inferred",
                        "receipt",
                        receipt["row_number"],
                        None,
                        "Позиция отсутствует в текущем snapshot; выбытие до загрузки определено автоматически.",
                    )
                )
            continue
        receipt_matches[receipt["row_number"]] = match
        receipt_title_by_source[match["source_id"]] = receipt["source_title"]
        if movement_status == "central_office_inventory_conflict":
            issues.append(
                Issue(
                    "warning",
                    "receipt_central_office_inventory_conflict",
                    "receipt",
                    receipt["row_number"],
                    match["source_id"],
                    "Строка помечена как ЦО, но товар присутствует в snapshot магазина.",
                )
            )
        if serial and serialized_identity(receipt["source_title"]) != serialized_identity(match["source_title"]):
            issues.append(
                Issue("blocker", "serialized_identity_conflict", "inventory", match["row_number"], match["source_id"],
                      f"Серийный номер {masked(serial)} связан с другим наименованием в поступлении.")
            )
            identity_by_source[match["source_id"]] = "conflict"
            receipt["match_status"] = "conflict"
        else:
            identity_by_source[match["source_id"]] = "matched" if serial else "not_applicable"
            receipt["match_status"] = "matched"

    for item in inventory:
        item["identity_status"] = identity_by_source.get(
            item["source_id"], "unmatched" if item["serial_full"] else "not_applicable"
        )
        combined_title = " ".join([item["source_title"], receipt_title_by_source.get(item["source_id"], "")])
        risks = risk_codes(combined_title, item["purchase_price"])
        for code in risks:
            severity = "warning" if code == "variant_ambiguous" else "blocker"
            issues.append(
                Issue(severity, code, "inventory", item["row_number"], item["source_id"],
                      "Нужна ручная проверка происхождения/варианта перед каталогом.")
            )
        item["risk_codes"] = risks
    return issues, receipt_matches


def decimal_json(value: Decimal) -> float:
    return float(value.quantize(Decimal("0.01")))


AUTO_BLOCK_CODES = {"serialized_identity_conflict", "replica_or_mimic", "authenticity_review"}


def review_state(existing: dict[str, Any] | None, item: dict[str, Any]) -> tuple[str, str]:
    hard_identity = item["identity_status"] == "conflict"
    hard_risk = any(
        code in {"replica_or_mimic", "authenticity_review"}
        for code in item.get("risk_codes", [])
    )
    override = bool(existing and existing.get("review_override") and text(existing.get("review_note")))
    authenticity = text(existing.get("authenticity_status")) if existing else "pending"
    eligibility = text(existing.get("eligibility_status")) if existing else "pending"

    if hard_identity:
        eligibility = "blocked"
    elif hard_risk and not override:
        authenticity = "blocked"
        eligibility = "blocked"
    elif override and authenticity in {"verified", "not_required"}:
        eligibility = "eligible"
    elif existing and eligibility == "blocked":
        previous_codes = {
            code.strip()
            for code in text(existing.get("block_reason")).split(",")
            if code.strip()
        }
        if previous_codes & AUTO_BLOCK_CODES:
            eligibility = "pending"
            if authenticity == "blocked":
                authenticity = "pending"

    return authenticity or "pending", eligibility or "pending"


def summarize(inventory: list[dict[str, Any]], receipts: list[dict[str, Any]], issues: list[Issue]) -> dict[str, Any]:
    inventory_cost = sum((row["purchase_price"] * row["quantity"] for row in inventory), Decimal("0"))
    inventory_retail = sum((row["retail_price"] * row["quantity"] for row in inventory), Decimal("0"))
    receipt_cost = sum((row["total_cost"] for row in receipts), Decimal("0"))
    receipt_retail = sum((row["total_price"] for row in receipts), Decimal("0"))
    margin = receipt_retail - receipt_cost
    movement_counts = {
        status: sum(row.get("movement_status") == status for row in receipts)
        for status in sorted({text(row.get("movement_status")) for row in receipts if row.get("movement_status")})
    }
    return {
        "valid": True,
        "inventory": {
            "rows": len(inventory),
            "units": sum(row["quantity"] for row in inventory),
            "purchase_value": decimal_json(inventory_cost),
            "retail_value": decimal_json(inventory_retail),
        },
        "receipts": {
            "rows": len(receipts),
            "units": sum(row["quantity"] for row in receipts),
            "purchase_value": decimal_json(receipt_cost),
            "retail_value": decimal_json(receipt_retail),
            "gross_profit": decimal_json(margin),
            "gross_margin": float(margin / receipt_retail) if receipt_retail else 0,
            "movement_counts": movement_counts,
            "central_office_units": sum(int(row.get("central_office_quantity") or 0) for row in receipts),
        },
        "issues": {
            "blockers": sum(issue.severity == "blocker" for issue in issues),
            "warnings": sum(issue.severity == "warning" for issue in issues),
            "by_code": {
                code: sum(issue.code == code for issue in issues)
                for code in sorted({issue.code for issue in issues})
            },
        },
    }


class Directus:
    def __init__(self, base_url: str, token: str) -> None:
        self.base_url = base_url.rstrip("/")
        self.token = token

    def request(self, method: str, path: str, payload: Any = None, params: dict[str, Any] | None = None) -> Any:
        query = f"?{urlencode(params)}" if params else ""
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8") if payload is not None else None
        request = Request(
            f"{self.base_url}{path}{query}",
            data=body,
            method=method,
            headers={"Authorization": f"Bearer {self.token}", "Content-Type": "application/json"},
        )
        try:
            with urlopen(request, timeout=60) as response:
                raw = response.read()
        except HTTPError as error:
            detail = error.read().decode("utf-8", errors="replace")[:500]
            raise RuntimeError(f"Directus {method} {path} failed: {error.code} {detail}") from error
        data = json.loads(raw.decode("utf-8")) if raw else {}
        return data.get("data", data)

    def first(self, collection: str, filters: dict[str, Any], fields: str = "*") -> dict[str, Any] | None:
        params: dict[str, Any] = {"fields": fields, "limit": 1}
        for key, value in filters.items():
            params[f"filter[{key}][_eq]"] = value
        rows = self.request("GET", f"/items/{collection}", params=params)
        return rows[0] if rows else None

    def upsert(self, collection: str, filters: dict[str, Any], payload: dict[str, Any]) -> dict[str, Any]:
        existing = self.first(collection, filters, "*")
        if existing:
            return self.request("PATCH", f"/items/{collection}/{quote(str(existing['id']), safe='')}", payload)
        return self.request("POST", f"/items/{collection}", payload)

    def delete_where(self, collection: str, field: str, value: str) -> None:
        rows = self.request("GET", f"/items/{collection}", params={f"filter[{field}][_eq]": value, "fields": "id", "limit": -1})
        for row in rows:
            self.request("DELETE", f"/items/{collection}/{quote(str(row['id']), safe='')}")

    def all(self, collection: str, filters: dict[str, Any], fields: str) -> list[dict[str, Any]]:
        params: dict[str, Any] = {"fields": fields, "limit": -1}
        for key, value in filters.items():
            params[f"filter[{key}][_eq]"] = value
        rows = self.request("GET", f"/items/{collection}", params=params)
        return rows if isinstance(rows, list) else []


def slug_id(title: str, source_id: str) -> str:
    ascii_text = unicodedata.normalize("NFKD", title).encode("ascii", "ignore").decode().lower()
    words = [word for word in re.sub(r"[^a-z0-9]+", "-", ascii_text).strip("-").split("-") if word]
    base = "-".join(words[:8]) or "product"
    return f"{base}-{source_id.replace('-', '')[:8]}"[:255]


def product_mapping(item: dict[str, Any]) -> tuple[str, str, str]:
    title = normalize(item["source_title"])
    group = normalize(item["source_group"])
    group_path = [
        normalize(segment)
        for segment in re.split(r"\\+|/|>", text(item.get("source_group_path")))
        if normalize(segment)
    ]
    group_leaf = group_path[-1] if group_path else ""
    brand = "apple" if any(term in title for term in ("apple", "iphone", "macbook", "airpods")) else (
        "samsung" if "samsung" in title else "xiaomi" if any(term in title for term in ("xiaomi", " mi ", "smartband")) else "other"
    )

    if "смарт очки" in group_path:
        return "device", "smart-electronics", brand
    if "смарт часы или браслеты" in group_path:
        return "device", "watches", brand
    if group_leaf in {"телефоны", "смартфоны"}:
        return "device", "smartphones", brand
    if group_leaf == "ноутбуки":
        return "device", "laptops", brand
    if group_leaf == "планшеты":
        return "device", "tablets", brand
    if group_leaf == "беспроводные наушники":
        return "device", "headphones", brand
    if group_leaf == "роутеры":
        return "device", "routers", brand
    if group_leaf == "смарт электроника":
        return "device", "smart-electronics", brand
    if group_leaf == "чехлы для смартфонов":
        return "accessory", "cases", brand
    if group_leaf == "защитные стекла":
        return "accessory", "protective-glass", brand
    if "зарядные устройства" in group_path:
        return "accessory", "chargers", brand
    if "ма ч" in group_leaf or "mah" in group_leaf:
        return "accessory", "power-banks", brand

    if group in {"телефоны", "смартфоны"}:
        return "device", "smartphones", brand
    if group == "ноутбуки":
        return "device", "laptops", brand
    if group == "планшеты":
        return "device", "tablets", brand
    if "час" in group:
        return "device", "watches", brand
    if "смарт очки" in group:
        return "device", "smart-electronics", brand
    if "науш" in group:
        return "device", "headphones", brand
    if "роут" in group:
        return "device", "routers", brand
    if "смарт электроника" in group:
        return "device", "smart-electronics", brand
    if "чех" in group:
        return "accessory", "cases", brand
    if "защит" in group:
        return "accessory", "protective-glass", brand
    if "кабел" in title:
        return "accessory", "cables", brand
    if "power bank" in title:
        return "accessory", "power-banks", brand
    if any(term in title for term in ("электробрит", "триммер")):
        return "device", "smart-electronics", brand
    if "заряд" in title or group in {"apple", "samsung"}:
        return "accessory", "chargers", brand
    return "accessory", "other-accessories", brand


def sync_eligible_product(client: Directus, item: dict[str, Any], stored: dict[str, Any], batch_name: str) -> str | None:
    if stored.get("eligibility_status") != "eligible" or not stored.get("review_override"):
        return None
    product_type, category_slug, brand_slug = product_mapping(item)
    category = client.first("product_categories", {"slug": category_slug}, "id")
    brand = client.first("product_brands", {"slug": brand_slug}, "id")
    if not category or not brand:
        raise RuntimeError(f"Missing catalog reference: {brand_slug}/{category_slug}")
    existing_product = client.first("products", {"source_system": item["source_system"], "source_id": item["source_id"]}, "*")
    product_id = str(existing_product["id"]) if existing_product else slug_id(item["source_title"], item["source_id"])
    payload = {
        "id": product_id,
        "sku": item["source_sku"],
        "status": "draft",
        "content_status": "needs_photo",
        "product_type": product_type,
        "condition": item["condition"],
        "sale_mode": "inquiry" if "variant_ambiguous" in item.get("risk_codes", []) else "reservation",
        "brand": brand["id"],
        "category": category["id"],
        "title": item["source_title"],
        "model": item["source_title"],
        "price": int(item["retail_price"].quantize(Decimal("1"))),
        "stock_quantity": item["quantity"],
        "stock_status": "available" if item["for_sale"] and item["quantity"] > 0 else "sold",
        "sort": 500,
        "source_system": item["source_system"],
        "source_id": item["source_id"],
        "import_batch": batch_name,
        "imported_at": datetime.now(timezone.utc).isoformat(),
        "admin_note": "Создано из приватного inventory snapshot. Публикация только после фото и контентного QA.",
    }
    if existing_product:
        inventory_payload = {
            key: payload[key]
            for key in (
                "sku",
                "price",
                "stock_quantity",
                "stock_status",
                "source_system",
                "source_id",
                "import_batch",
                "imported_at",
            )
        }
        product = client.request("PATCH", f"/items/products/{quote(product_id, safe='')}", inventory_payload)
    else:
        product = client.request("POST", "/items/products", payload)
    if product_type == "device":
        client.upsert("device_details", {"product": product_id}, {"product": product_id, "serial": masked(item["serial_full"])})
    else:
        existing_details = client.first("accessory_details", {"product": product_id}, "id")
        if not existing_details:
            client.upsert(
                "accessory_details",
                {"product": product_id},
                {"product": product_id, "compatibility_mode": "universal", "specifications": {}},
            )
    existing_listing = client.first(
        "product_channel_listings", {"product": product_id, "channel": "avito"}, "*"
    )
    listing_payload = {"external_id": f"isvoi-{item['source_id']}"}
    if existing_listing:
        client.request(
            "PATCH",
            f"/items/product_channel_listings/{quote(str(existing_listing['id']), safe='')}",
            listing_payload,
        )
    else:
        client.request(
            "POST",
            "/items/product_channel_listings",
            {"product": product_id, "channel": "avito", "status": "draft", **listing_payload},
        )
    return str(product.get("id", product_id))


def archive_previous_inventory_batches(
    client: Directus, batch_id: str, batch_name: str, source_system: str
) -> int:
    archived = 0
    completed_statuses = {"checked", "applied", "applied_with_blocks", "failed"}
    batches = client.all(
        "inventory_import_batches",
        {"source_system": source_system},
        "id,status",
    )
    for batch in batches:
        previous_id = str(batch.get("id") or "")
        if previous_id == batch_id or batch.get("status") not in completed_statuses:
            continue
        client.request(
            "PATCH",
            f"/items/inventory_import_batches/{quote(previous_id, safe='')}",
            {"status": "archived"},
        )
        open_issues = client.all(
            "inventory_import_issues",
            {"batch": previous_id, "resolved": False},
            "id",
        )
        for issue in open_issues:
            client.request(
                "PATCH",
                f"/items/inventory_import_issues/{quote(str(issue['id']), safe='')}",
                {
                    "resolved": True,
                    "resolution_note": f"Архивировано после применения партии {batch_name}.",
                },
            )
        archived += 1
    return archived


def apply_snapshot(
    client: Directus,
    batch_id: str,
    batch_name: str,
    source_system: str,
    inventory: list[dict[str, Any]],
    receipts: list[dict[str, Any]],
    issues: list[Issue],
    receipt_matches: dict[int, dict[str, Any]],
    missing_items: list[dict[str, Any]],
    confirm_missing_deactivation: bool,
) -> dict[str, int]:
    issue_codes_by_source: dict[str, set[str]] = {}
    for issue in issues:
        if issue.source_id:
            issue_codes_by_source.setdefault(issue.source_id, set()).add(issue.code)

    stored_by_source: dict[str, dict[str, Any]] = {}
    products_synced = 0
    for item in inventory:
        item["source_system"] = source_system
        existing = client.first("inventory_items", {"source_system": source_system, "source_id": item["source_id"]}, "*")
        authenticity, eligibility = review_state(existing, item)
        payload = {
            key: (decimal_json(value) if isinstance(value, Decimal) else value)
            for key, value in item.items()
            if key not in {"row_number", "risk_codes", "source_system"}
        }
        payload.update(
            {
                "source_system": source_system,
                "identity_status": item["identity_status"],
                "authenticity_status": authenticity or "pending",
                "eligibility_status": eligibility or "pending",
                "block_reason": ", ".join(sorted(issue_codes_by_source.get(item["source_id"], set()))) or None,
                "last_seen_batch": batch_id,
            }
        )
        stored = client.upsert("inventory_items", {"source_system": source_system, "source_id": item["source_id"]}, payload)
        item_id = str(stored["id"])
        stored_by_source[item["source_id"]] = stored
        item["stored_id"] = item_id
        product_id = sync_eligible_product(client, item, stored, batch_name)
        if product_id:
            products_synced += 1
            client.request("PATCH", f"/items/inventory_items/{quote(item_id, safe='')}", {"product": product_id})

    deactivated = 0
    if confirm_missing_deactivation:
        for missing in missing_items:
            client.request(
                "PATCH",
                f"/items/inventory_items/{quote(str(missing['id']), safe='')}",
                {
                    "quantity": 0,
                    "for_sale": False,
                    "eligibility_status": "pending",
                    "block_reason": "missing_from_snapshot",
                },
            )
            product = missing.get("product")
            product_id = product.get("id") if isinstance(product, dict) else product
            if product_id:
                client.request(
                    "PATCH",
                    f"/items/products/{quote(str(product_id), safe='')}",
                    {"stock_quantity": 0, "stock_status": "sold"},
                )
                listings = client.all(
                    "product_channel_listings",
                    {"product": product_id},
                    "id,status",
                )
                for listing in listings:
                    if listing.get("status") == "active":
                        client.request(
                            "PATCH",
                            f"/items/product_channel_listings/{quote(str(listing['id']), safe='')}",
                            {"status": "draft", "sync_status": "inventory_missing"},
                        )
            deactivated += 1

    client.delete_where("inventory_receipt_lines", "batch", batch_id)
    for receipt in receipts:
        matched = receipt_matches.get(receipt["row_number"])
        stored = stored_by_source.get(matched["source_id"]) if matched else None
        payload = {
            key: (decimal_json(value) if isinstance(value, Decimal) else value)
            for key, value in receipt.items()
            if key not in {"match_status", "match_note"}
        }
        payload.update(
            {
                "batch": batch_id,
                "inventory_item": stored.get("id") if stored else None,
                "product": stored.get("product") if stored else None,
                "match_status": receipt.get("match_status", "unmatched"),
                "match_note": receipt.get("match_note"),
            }
        )
        client.request("POST", "/items/inventory_receipt_lines", payload)

    client.delete_where("inventory_import_issues", "batch", batch_id)
    for issue in issues:
        linked_item = stored_by_source.get(issue.source_id or "")
        client.request(
            "POST",
            "/items/inventory_import_issues",
            {
                "batch": batch_id,
                **asdict(issue),
                "inventory_item": linked_item.get("id") if linked_item else None,
            },
        )
    batches_archived = archive_previous_inventory_batches(
        client, batch_id, batch_name, source_system
    )
    return {
        "inventory_items": len(inventory),
        "receipt_lines": len(receipts),
        "issues": len(issues),
        "products_synced": products_synced,
        "missing_items": len(missing_items),
        "items_deactivated": deactivated,
        "batches_archived": batches_archived,
    }


def find_missing_items(
    client: Directus,
    source_system: str,
    inventory: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    current_ids = {item["source_id"] for item in inventory}
    existing = client.all(
        "inventory_items",
        {"source_system": source_system},
        "id,source_id,source_sku,source_title,quantity,product",
    )
    return [item for item in existing if item.get("source_id") not in current_ids]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--inventory", required=True)
    parser.add_argument("--receipts")
    parser.add_argument("--batch", required=True)
    parser.add_argument("--batch-id")
    parser.add_argument("--source-system", default="store_inventory")
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--confirm-missing-deactivation", action="store_true")
    args = parser.parse_args()

    inventory = parse_inventory(Path(args.inventory))
    receipts = parse_receipts(Path(args.receipts)) if args.receipts else []
    issues, receipt_matches = reconcile(inventory, receipts)
    base_url = os.getenv("DIRECTUS_URL", "")
    token = os.getenv("INVENTORY_IMPORT_DIRECTUS_TOKEN") or os.getenv("DIRECTUS_TOKEN", "")
    client = Directus(base_url, token) if base_url and token else None
    missing_items = find_missing_items(client, args.source_system, inventory) if client else []
    for missing in missing_items:
        issues.append(
            Issue(
                severity="warning",
                code="missing_from_snapshot",
                source_kind="inventory",
                row_number=None,
                source_id=text(missing.get("source_id")) or None,
                message=(
                    f"Позиция {text(missing.get('source_sku')) or 'без кода'} отсутствует в новом snapshot; "
                    "остаток не будет обнулён без отдельного подтверждения."
                ),
            )
        )
    summary = summarize(inventory, receipts, issues)
    summary["missing_from_snapshot"] = {
        "count": len(missing_items),
        "deactivation_confirmed": args.confirm_missing_deactivation,
    }
    summary["issue_samples"] = [
        {**asdict(issue), "source_id": issue.source_id[:8] if issue.source_id else None}
        for issue in issues[:40]
    ]

    if args.apply:
        if not args.batch_id:
            raise RuntimeError("--batch-id is required with --apply")
        if not client:
            raise RuntimeError("DIRECTUS_URL and INVENTORY_IMPORT_DIRECTUS_TOKEN are required")
        summary["applied"] = apply_snapshot(
            client, args.batch_id, args.batch, args.source_system,
            inventory, receipts, issues, receipt_matches, missing_items,
            args.confirm_missing_deactivation,
        )

    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
